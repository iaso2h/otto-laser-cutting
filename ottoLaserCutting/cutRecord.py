import util
from util import pr
from config import cfg

import keySet

import shutil
import datetime
import time
import os
import re
import numpy
import win32api, win32con, win32gui, win32process
import psutil
import easyocr
import json
from PIL import Image, ImageFilter, ImageGrab
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from typing import Optional
from pathlib import Path

SCREENSHOT_DIR_PATH = Path(cfg.paths.otto, r"存档/截图")
CUT_RECORD_PATH     = Path(cfg.paths.otto, r"存档/开料记录.xlsx")
LASER_OCR_FIX_PATH  = Path(cfg.paths.otto, r"辅助程序/激光名称OCR修复规则.json")
MESSAGEBOX_TITLE = "激光开料"
pr = util.pr

def getWorkbook() -> Workbook:
    """
    Loads or creates an Excel workbook for cut records.

    Returns:
        openpyxl.Workbook: Existing workbook if CUT_RECORD_PATH exists,
                           otherwise a new Workbook instance.
    """
    if CUT_RECORD_PATH.exists():
        return load_workbook(str(CUT_RECORD_PATH))
    else:
        return Workbook()


screenshotPaths = []

def findMessageBoxWindow() -> Optional[int]:
    print("Finding prompted window")
    startTime = time.time()
    timeout = 5  # seconds
    while time.time() - startTime < timeout:
        hwnd = win32gui.FindWindow(None, MESSAGEBOX_TITLE)
        if hwnd != 0:
            return hwnd
        time.sleep(1)  # Poll every 1 seconds
    return None


def initSheetFromScreenshots(wb: Workbook) -> None:  # {{{
    """
    Initializes workbook sheets from screenshot files.

    Scans the screenshot directory for PNG files with specific dimensions (1080x1920).
    For each unique year-month prefix found in screenshot filenames, creates a new sheet
    in the workbook with standard headers if it doesn't already exist.

    Args:
        wb (Workbook): The Excel workbook to initialize sheets in.

    The created sheets will have columns for:
    - Layout file
    - Completion time
    - Order number
    - Model (quantity)
    - Cut/required quantity
    - Screenshot file
    """
    yearMonthPrefix = []
    sheetNames = wb.sheetnames
    for p in SCREENSHOT_DIR_PATH.iterdir():
        if p.suffix == ".png":
            with Image.open(p) as img:
                width, height = img.size
                if width != 1080 or height != 1920:
                    continue

            screenshotPaths.append(p)
            dateStamp = p.stem[5:12]
            if dateStamp not in yearMonthPrefix:
                yearMonthPrefix.append(dateStamp)

    for n in yearMonthPrefix:
        if n not in sheetNames:
            ws = wb.create_sheet(n, 0)
            ws["A1"].value = "排样文件"
            ws["B1"].value = "完成时间"
            ws["C1"].value = "单号"
            ws["D1"].value = "型号(数量)"
            ws["E1"].value = "已切量/需求量"
            ws["F1"].value = "截图文件" # }}}


def takeScreenshot(screenshot: Optional[Image.Image] = None) -> None:  # {{{
    """
    Takes a screenshot and records cutting information in an Excel file.

    This function handles screenshot capture and logging of laser cutting operations:
    1. Checks for modifier keys (Ctrl/Shift) to trigger alternative actions
    2. Identifies the active TubePro window to get part filename
    3. Captures screen if no image is provided
    4. Saves screenshot and records metadata (timestamp, filename) in Excel
    5. Shows success notification

    Args:
        screenshot: Optional pre-captured image to use instead of grabbing new screenshot

    Returns:
        None: Opens the record file or performs relinking based on key modifiers,
              otherwise shows success message

    Note:
        - Requires TubePro.exe to be running for normal operation
        - Maintains Excel records with specific columns (A1-F1 headers)
        - Copies records to shared location when not running on OT03 machine
    """
    if "ctrl" in keySet.keys:
        return os.startfile(CUT_RECORD_PATH)
    elif "shfit" in keySet.keys:
        return relinkScreenshots()

    # Get laser file info
    hwndTitles = {}
    def winEnumHandler(hwnd, ctx):
        if win32gui.IsWindowVisible(hwnd):
            windowText = win32gui.GetWindowText(hwnd)
            if windowText:
                hwndTitles[hwnd] = windowText
        return True

    win32gui.EnumWindows(winEnumHandler, None)

    partFileName = ""
    for hwnd, title in hwndTitles.items():
        if title.startswith("TubePro"):
            _, pId = win32process.GetWindowThreadProcessId(hwnd)
            pName = psutil.Process(pId).name()
            if pName == "TubePro.exe":
                partFileName = re.sub(r"^TubePro(\(.+?\))? (.+\.zzx).*?$", r"\2", title, re.IGNORECASE)

                if win32gui.IsIconic(hwnd):
                    win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                win32gui.SetForegroundWindow(hwnd)
                break

    if not partFileName:
        return pr("Screenshot taking is abort due to TubePro is not running.")

    if not screenshot:
        screenshot = ImageGrab.grab()

    # Check current foreground program
    datetimeNow = datetime.datetime.now()
    excelTimeStamp = datetimeNow.strftime("%Y/%m/%d %H:%M:%S")
    screenshotPath = util.screenshotSave(screenshot, "屏幕截图", SCREENSHOT_DIR_PATH)

    # Using OCR to get process count
    wb = getWorkbook()
    sheetName = screenshotPath.stem[5:12]
    try:
        ws = wb[sheetName]
    except Exception:
        ws = wb.create_sheet(sheetName, 0)
        ws["A1"].value = "排样文件"
        ws["B1"].value = "完成时间"
        ws["C1"].value = "单号"
        ws["D1"].value = "型号(数量)"
        ws["E1"].value = "已切量/需求量"
        ws["F1"].value = "截图文件"

    newRecord(ws, screenshotPath, partFileName, excelTimeStamp)
    savePath = util.saveWorkbook(wb, CUT_RECORD_PATH)

    if os.getlogin() != "OT03":
        shutil.copy2(savePath, Path(SCREENSHOT_DIR_PATH, "开料记录.xlsx"))

    win32api.MessageBox(
                None,
                f"记录成功",
                MESSAGEBOX_TITLE,
                4096 + 64 + 0
            )
    #   MB_SYSTEMMODAL==4096
    ##  Button Styles:
    ### 0:OK  --  1:OK|Cancel -- 2:Abort|Retry|Ignore -- 3:Yes|No|Cancel -- 4:Yes|No -- 5:Retry|No -- 6:Cancel|Try Again|Continue
    ##  To also change icon, add these values to previous number
    ### 16 Stop-sign  ### 32 Question-mark  ### 48 Exclamation-point  ### 64 Information-sign ('i' in a circle)
# }}}


def getImgInfo(p: Path) -> None:  # {{{
    """
    Extracts and processes text information from an image file using OCR.

    Args:
        p (Path): Path to the image file to process.

    Returns:
        tuple: A 3-tuple containing:
            - partFileName (str): Extracted and cleaned filename from image title
            - partProcessCount (str): Extracted process count from image
            - timeStamp (str): Extracted and formatted timestamp from image

    The function performs the following operations:
    1. Crops specific regions of interest from the image (title, process count, timestamp)
    2. Checks for completion status via pixel color detection
    3. Uses EasyOCR to extract text from image regions
    4. Applies text cleaning and pattern substitutions
    5. Handles different timestamp formats based on completion status
    6. Removes illegal characters from all extracted text fields
    """
    reader = easyocr.Reader(["ch_sim", "en"])

    with Image.open(p) as img:
        imgTitle        = img.crop((91, 0, 900, 25))
        imgProcessCount = img.crop((550, 1665, 765, 1685))
        cvTitle = numpy.array(imgTitle)[:, :, ::-1].copy()
        cvProcessCount = numpy.array(imgProcessCount)[:, :, ::-1].copy()

        imgRGB = img.convert("RGB")
        targetCompletedPixel = imgRGB.getpixel((15, 1810))
        if targetCompletedPixel == (170, 170, 0) or targetCompletedPixel == (255, 155, 155):
            # Also treat A21 error code as completion message
            targetCompletedChk = True
        else:
            targetCompletedChk = False

        if targetCompletedChk:
            imgTimeStamp = img.crop((104, 1777, 240, 1792))
            cvTimeStamp  = numpy.array(imgTimeStamp)[:, :, ::-1].copy()
        else:
            imgTimeStamp = img.crop((91, 1755, 185, 1864)).filter(ImageFilter.EDGE_ENHANCE)
            cvTimeStamp  = numpy.array(imgTimeStamp)[:, :, ::-1].copy()

    titleRead = reader.readtext(cvTitle)
    processCountRead = reader.readtext(cvProcessCount)
    timeStampRead = reader.readtext(cvTimeStamp)
    partFileName = ""
    partProcessCount = ""
    timeStamp = p.stem[5:] # Default time stamp
    if titleRead:
        for text in titleRead:
            partFileName = partFileName + " " + text[1]
            suffixMatch = re.search(r"\.zzx", partFileName, flags=re.IGNORECASE)
            if suffixMatch:
                partFileName = partFileName[:suffixMatch.span()[1]]
            partFileName = partFileName.strip()
            with open(LASER_OCR_FIX_PATH, "r", encoding="utf-8") as pat:
                commonFix = json.load(pat)
            for key, val in commonFix.items():
                pattern = re.compile(key, re.IGNORECASE)
                partFileName = pattern.sub(val, partFileName)

    if processCountRead:
        if len(processCountRead) == 2:
            # In case recognition result is 2
            partProcessCount = processCountRead[1][1]

    if timeStampRead:
        timeStamp = timeStampRead[len(timeStampRead) - 1][1]
        if not targetCompletedChk:
            timeStamp = p.stem[5:9] + "/" + timeStamp # Add year prefix

        commonFix = {
                "l": "1",
                "i": "1",
                ";": ":",
                ".": ":",
                ",": ":",
                "+": ":",
                }
        for key, val in commonFix.items():
            timeStamp = timeStamp.replace(key, val)

    partFileName     = ILLEGAL_CHARACTERS_RE.sub("", partFileName)
    timeStamp        = ILLEGAL_CHARACTERS_RE.sub("", timeStamp)
    partProcessCount = ILLEGAL_CHARACTERS_RE.sub("", partProcessCount)
    return partFileName, partProcessCount, timeStamp # }}}


def validScreenshotPath(cell):  # {{{
    """
    Check if a cell contains a valid screenshot file path.

    Args:
        cell: The cell object to validate, expected to have a 'value' attribute.

    Returns:
        bool: True if the cell value is a string and points to an existing file, False otherwise.
    """
    if (
        not cell.value
        or not isinstance(cell.value, str)
        or not Path(cell.value).exists()
    ):
        return False
    else:
        return True # }}}


def newRecord(ws: Worksheet, p: str, partFileName: Optional[str]=None, timeStamp: Optional[str]=None):
    """
    Creates a new record in the worksheet with part processing information.

    Args:
        ws (Worksheet): The worksheet to add the record to
        p (str): Path to the image file containing processing data
        partFileName (str, optional): Name of the part file. If not provided, extracted from image.
        timeStamp (str, optional): Timestamp of processing. If not provided, extracted from image.

    The function either uses provided partFileName/timeStamp or extracts them from the image.
    Extracts process count from image using OCR when needed. Adds a new row with:
    - Part filename (column A)
    - Timestamp (column B, formatted)
    - Process count (column E, as text)
    - Hyperlink to image (column F)
    """
    if not partFileName or not timeStamp:
        partFileName, partProcessCount, timeStamp = getImgInfo(p)
    else:
        reader = easyocr.Reader(["en"])
        partProcessCount = ""
        with Image.open(p) as img:
            imgProcessCount = img.crop((550, 1665, 765, 1685))
            cvProcessCount = numpy.array(imgProcessCount)[:, :, ::-1].copy()
            processCountRead = reader.readtext(cvProcessCount)
            if processCountRead:
                if len(processCountRead) == 2:
                    # In case recognition result is 2
                    partProcessCount = processCountRead[1][1]
                    partProcessCount = ILLEGAL_CHARACTERS_RE.sub("", partProcessCount)

    rowNew = ws.max_row + 1
    ws[f"A{rowNew}"].value = partFileName
    ws[f"B{rowNew}"].value = timeStamp
    ws[f"B{rowNew}"].number_format = "yyyy/m/d h:mm:ss"
    ws[f"E{rowNew}"].value = str(partProcessCount)
    ws[f"E{rowNew}"].number_format = "@"
    ws[f"F{rowNew}"].hyperlink = str(p)


def updateScreenshotRecords():  # {{{
    """
    Updates the screenshot records in the workbook by comparing timestamps.
    For each screenshot path, checks if it's newer than the last recorded screenshot
    in the corresponding worksheet. If newer or if worksheet is empty, adds a new record.
    Saves the updated workbook to CUT_RECORD_PATH.

    Args:
        None (uses module-level variables: screenshotPaths, CUT_RECORD_PATH)

    Effects:
        Modifies the workbook by adding new records when appropriate
        Saves the workbook to CUT_RECORD_PATH
    """
    wb = getWorkbook()
    initSheetFromScreenshots(wb)
    for p in screenshotPaths:
        sheetName = p.stem[5:12]
        ws = wb[sheetName]
        rowMax = ws.max_row
        # fix rowMax to row that contain valid screenshot path
        lastDatetime = None
        if rowMax != 1:
            # Get the valid last datetime
            while rowMax > 1:
                lastScreenshotCell = ws[f"F{rowMax}"]
                if not validScreenshotPath(lastScreenshotCell):
                    rowMax = rowMax - 1
                    continue
                if "\n" in str(lastScreenshotCell.value).strip():
                    paths = str(lastScreenshotCell.value).strip().split("\n")
                    lastPath = Path(paths[len(paths) - 1])
                else:
                    lastPath = Path(lastScreenshotCell.value)

                try:
                    lastDatetime = datetime.datetime.strptime(str(lastPath.stem)[5:], "%Y-%m-%d %H%M%S")
                    break
                except ValueError:
                    rowMax = rowMax - 1
                    continue

            if not lastDatetime:
                newRecord(ws, p)
            else:
                currentDatetime = datetime.datetime.strptime(str(p.stem)[5:], "%Y-%m-%d %H%M%S")
                # Only save screenshots that are newer than the last one
                if lastDatetime < currentDatetime:
                    newRecord(ws, p)
        else:
            # Start in a new worksheet
            newRecord(ws, p)

    util.saveWorkbook(wb, CUT_RECORD_PATH) # }}}


def relinkScreenshots():
    """
    Relinks screenshot hyperlinks in the cut record workbook.

    This function checks each cell in the worksheet for valid screenshot paths and updates
    the hyperlinks in column F. It skips empty rows and invalid paths. When multiple paths
    are found in a cell (separated by newlines), it uses the last one. Only existing .png
    files are processed.

    Note: Requires 'ctrl' key in keySet to execute the file opening operation.
    """
    if "ctrl" in keySet.keys:
        return os.startfile(CUT_RECORD_PATH)
    # TODO: highlight invalid ones
    wb = getWorkbook()
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
            for cell in row:
                if not validScreenshotPath(cell):
                    continue

                if "\n" in str(cell.value).strip():
                    screenshotPaths = str(cell.value).strip().split("\n")
                    screenshotPath = Path(screenshotPaths[len(screenshotPaths) - 1])
                else:
                    screenshotPath = Path(str(cell.value))

                if screenshotPath.exists() and screenshotPath.suffix == ".png":
                    ws[f"F{cell.row}"].hyperlink = cell.value

    util.saveWorkbook(wb, CUT_RECORD_PATH)
