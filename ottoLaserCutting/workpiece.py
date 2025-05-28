import config
import util
import keySet
import subprocess
from config import cfg

import re
import os
import json
import datetime
import win32api, win32con
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from pathlib import Path
from openpyxl.styles import Font
from openpyxl.styles.numbers import BUILTIN_FORMATS
# https://openpyxl.readthedocs.io/en/3.1.3/_modules/openpyxl/styles/numbers.html
from decimal import Decimal

pr = util.pr
WORKPIECE_INFO_PATH = Path(cfg.paths.otto, r"存档/零件规格总览.xlsx")
WORKPIECE_DICT = Path(cfg.paths.otto, r"辅助程序/workpieceDict.json")


def bankRound(precision: float, digitLiteral: str) -> float:
    """
    Rounds a decimal number using banker's rounding (ROUND_HALF_UP) to the specified precision.

    Args:
        precision: The decimal precision to round to (e.g. 0.01 for 2 decimal places)
        digitLiteral: The number to round, provided as a string to avoid floating-point precision issues

    Returns:
        The rounded number as a float
    """
    return float(
        Decimal(digitLiteral).quantize(Decimal(precision), rounding="ROUND_HALF_UP")
    )


def removeDummyLaserFile(p: Path) -> None:
    """
    Removes a dummy laser file if it meets specific criteria.

    Args:
        p (Path): Path object representing the file to be checked and removed.

    The function checks if the file has no suffix and zero size. If both conditions
    are met, it attempts to remove the file silently (ignores any removal errors).
    """
    if p.suffix == "" and p.stat().st_size == 0:
        try:
            os.remove(p)
        except:
            pass


def workpieceNamingVerification():
    """
    Verifies laser cutting workpiece file names against naming conventions.
    Checks all files in LASER_FILE_DIR_PATH:
    1. If Ctrl key is pressed, opens file explorer at the directory
    2. Otherwise scans for .zx/.zzx files and validates their names against cfg.patterns.laserFile
    3. Prints either validation results or "All files match" message
    Returns: subprocess.Popen object if Ctrl pressed, None otherwise
    """
    if "ctrl" in keySet.keys:
        return subprocess.Popen(rf'explorer /select, "{config.LASER_FILE_DIR_PATH}"')
    laserFilePaths = util.getAllLaserFiles()
    if not laserFilePaths:
        pr("All files match the naming convention!")
    invalidFilePathFoundChk = False
    for _, p in enumerate(laserFilePaths):
        if p.suffix == ".zx" or p.suffix == ".zzx":
            fileNameMatch = cfg.patterns.laserFile.match(
                str(p.stem)
            )
            if not fileNameMatch:
                invalidFilePathFoundChk = True
                pr(f'------------------------\n({_}): "{p.stem}"')
    if not invalidFilePathFoundChk:
        pr("没有不规范的工件名称")


def removeRedundantLaserFile() -> None:
    """
    Removes redundant laser files from the configured directory.

    This function scans the LASER_FILE_DIR_PATH for .zx files that have a corresponding
    .zzx file with a newer modification time. Such files are considered redundant and
    are deleted. If the 'ctrl' key is pressed, it opens the directory in Explorer instead.

    Displays a summary of deleted files via console and a system message box.
    Returns None.

    Behavior:
    - Skips files containing 'demo' in their name (case-insensitive)
    - Only processes files when LASER_FILE_DIR_PATH exists
    - Shows deletion count and list of deleted files
    - Uses MB_SYSTEMMODAL (4096) + MB_ICONINFORMATION (64) for the message box
    """
    if "ctrl" in keySet.keys:
        return subprocess.Popen(rf'explorer /select, "{config.LASER_FILE_DIR_PATH}"')
    rawLaserFile = []

    if not config.LASER_FILE_DIR_PATH.exists():
        return

    for p in config.LASER_FILE_DIR_PATH.iterdir():
        p = util.strStandarize(p)
        if p.is_file() and "demo" not in p.stem.lower():
            rawLaserFile.append(p)

    pDeletedStr = []
    for p in rawLaserFile:
        laserFile = Path(p.parent, p.stem + ".zzx")
        if laserFile.exists() and laserFile.stat().st_mtime > p.stat().st_mtime:
            try:
                os.remove(p)
                pDeletedStr.append(str(p))
            except:
                pass

    if len(pDeletedStr) > 0:
        pr(f"{len(pDeletedStr)} redundant .zx files has been deleted:")
        for pStr in pDeletedStr:
            pr(pStr)
        win32api.MessageBox(
                    None,
                    f"{len(pDeletedStr)}个冗余文件已经被删除",
                    "Info",
                    4096 + 64 + 0
                )
        #   MB_SYSTEMMODAL==4096
        ##  Button Styles:
        ### 0:OK  --  1:OK|Cancel -- 2:Abort|Retry|Ignore -- 3:Yes|No|Cancel -- 4:Yes|No -- 5:Retry|No -- 6:Cancel|Try Again|Continue
        ##  To also change icon, add these values to previous number
        ### 16 Stop-sign  ### 32 Question-mark  ### 48 Exclamation-point  ### 64 Information-sign ('i' in a circle)
    else:
        pr("No redundant .zx files")


def exportDimensions():
    """
    Exports workpiece dimensions to Excel files with formatting and calculations.

    Generates two Excel files:
    1. WORKPIECE_INFO_PATH: Detailed workpiece specifications
    2. "零件规格总览.xlsx": Summary overview in warehousing directory

    Features:
    - Automatic dimension calculations
    - Formatting for laser-cut parts
    - Area overrides from workpieceDict
    - Table formatting with striped rows
    - Automatic timestamping
    - Protection against duplicate entries

    Handles special cases:
    - Laser cutting machine files (.zx/.zzx)
    - Welding combinations
    - Various tube types (main tube, handle tube, etc.)
    - Area calculation overrides
    """
    dstPath1 = WORKPIECE_INFO_PATH
    dstPath2 = Path(cfg.paths.warehousing, "零件规格总览.xlsx")
    boldFont = Font(bold=True)
    if "ctrl" in keySet.keys:
        return os.startfile(dstPath1)
    laserFilePaths = util.getAllLaserFiles()
    with open(WORKPIECE_DICT, "r", encoding="utf-8") as f:
        workpieceDict = json.load(f)

    wb = Workbook()
    ws = wb["Sheet"]
    ws["A1"] = "更新时间:" + str(datetime.datetime.now().strftime("%Y-%m-%d %H%M%S%f"))
    ws.merge_cells("B1:F1")
    ws["A2"].value = "零件名称"
    ws.column_dimensions["A"].width = 25
    ws["B2"].value = "外发别名"
    ws.column_dimensions["B"].width = 14
    ws["C2"].value = "规格"
    ws.column_dimensions["C"].width = 20
    ws["D2"].value = "材料"
    ws.column_dimensions["D"].width = 9
    ws["E2"].value = "参数一"
    ws.column_dimensions["E"].width = 8
    ws["F2"].value = "参数二"
    ws.column_dimensions["F"].width = 8
    ws["G2"].value = "长度"
    ws.column_dimensions["G"].width = 8
    ws["H2"].value = "方数(m²)"
    ws.column_dimensions["H"].width = 9.5
    ws["I2"].value = "焊接散件"
    ws.column_dimensions["I"].width = 12
    workpieceFullNamesWithDimension = []
    workpieceNickNames = workpieceDict["nickname"]
    # <fullPartName>: ["<nickName>", "<comment>"]
    for lIdx, p in enumerate(laserFilePaths):
        rowMax = ws.max_row + 1
        if p.suffix == ".zx" or p.suffix == ".zzx":
            fileNameMatch = cfg.patterns.laserFile.match(str(p.stem))
            boldFontChk = True
        else:
            fileNameMatch = cfg.patterns.laserFile.match(str(p.name))
            boldFontChk = False

        workpieceNickName  = ""
        workpieceDimension = ""
        surfaceAreaEval = 0
        fileNameMatchTick = False

        if not fileNameMatch:
            if p.suffix == ".zx" or p.suffix == ".zzx":
                workpieceFullName = p.stem
            else:
                workpieceFullName = p.name

            if workpieceFullName in workpieceFullNamesWithDimension:
                removeDummyLaserFile(p)
                continue
            else:
                workpieceFullNamesWithDimension.append(workpieceFullName)

            # Set font to be bold when the workpiece is produce in lasercutting machine
            if (
                p.suffix == ".zx"
                or p.suffix == ".zzx"
                or "焊接组合" in p.stem
                or "弯管" in p.stem
                or "主体管" in p.stem
                or "把手管" in p.stem
                or "辅助轮管" in p.stem
                or "调节管" in p.stem
                or "侧管" in p.stem
                or "支管" in p.stem
                or "座架" in p.stem
                or "扶手管" in p.stem
                or "铝拐臂" in p.stem
            ):
                for colIdx in range(1, 10):
                    ws.cell(row=rowMax, column=colIdx).font = boldFont

            ws[f"A{rowMax}"].value = workpieceFullName
            ws[f"A{rowMax}"].number_format = "@"
            # namingly ws[f"A{rowMax}"].number_format = BUILTIN_FORMATS[49]

            if workpieceFullName.endswith(" 焊接组合"):
                workpieceNickName = workpieceFullName.replace(" 焊接组合", "")
            if workpieceFullName in workpieceNickNames:
                workpieceNickName = workpieceNickNames[workpieceFullName][0]
                if workpieceNickNames[workpieceFullName][1]:
                    comment = Comment(workpieceNickNames[workpieceFullName][1], "阮焕")
                    comment.width = 300
                    comment.height = 150
                    ws[f"B{rowMax}"].comment = comment

            ws[f"B{rowMax}"].value = workpieceNickName
            ws[f"B{rowMax}"].number_format = "@"

        else:
            fileNameMatchTick = True

            productId          = fileNameMatch.group(1)
            productIdNote      = fileNameMatch.group(2) # name
            workpieceName           = fileNameMatch.group(3)
            if workpieceName and "飞切" in workpieceName:
                workpieceName = re.sub(r"[有无]飞切", "", workpieceName)
                workpieceName = workpieceName.replace("()", "")
            workpieceComponentName  = fileNameMatch.group(4) # Optional
            workpieceMaterial       = fileNameMatch.group(5)
            workpieceDimension = fileNameMatch.group(7)
            if workpieceDimension:
                workpieceDimension = workpieceDimension.replace("_", "*")
                workpieceDimension = workpieceDimension.replace("x", "*")
                workpieceDimension = util.diametartSymbolUnify(workpieceDimension)
                workpieceDimension = workpieceDimension.strip()
            else:
                workpieceDimension = ""

            workpiece1stParameter = fileNameMatch.group(8)
            workpiece2ndParameter = fileNameMatch.group(9) # Optional

            # DEPRECATED:
            # workpiece2ndParameterNum = fileNameMatch.group(11) # Optional
            workpieceLength = fileNameMatch.group(12)

            workpieceFullName = "{} {}".format(productId, workpieceName)
            workpieceFullNameWithDimension = "{} {}".format(workpieceFullName, workpieceDimension)
            if workpieceFullNameWithDimension in workpieceFullNamesWithDimension:
                removeDummyLaserFile(p)
                continue
            else:
                workpieceFullNamesWithDimension.append(workpieceFullNameWithDimension)

            tailingWorkpiece = fileNameMatch.group(14)        # Optional
            workpieceLongTubeLength = fileNameMatch.group(16) # Optional

            # Set font to be bold when the workpiece is produce in lasercutting machine
            if (
                p.suffix == ".zx"
                or p.suffix == ".zzx"
                or "焊接组合" in p.stem
                or "弯管" in p.stem
                or "主体管" in p.stem
                or "把手管" in p.stem
                or "辅助轮管" in p.stem
                or "调节管" in p.stem
                or "侧管" in p.stem
                or "支管" in p.stem
                or "座架" in p.stem
                or "扶手管" in p.stem
                or "铝拐臂" in p.stem
            ):
                for colIdx in range(1, 10):
                    ws.cell(row=rowMax, column=colIdx).font = boldFont

            ws[f"A{rowMax}"].value = workpieceFullName
            ws[f"A{rowMax}"].number_format = "@"
            # namingly ws[f"A{rowMax}"].number_format = BUILTIN_FORMATS[49]
            if workpieceFullName in workpieceNickNames:
                workpieceNickName = workpieceNickNames[workpieceFullName][0]
                comment = workpieceNickNames[workpieceFullName][1]
                if comment:
                    ws[f"B{rowMax}"].comment = Comment(comment, "阮焕")
                    ws[f"B{rowMax}"].comment.width = 300
                    ws[f"B{rowMax}"].comment.height = 150
            ws[f"B{rowMax}"].value = workpieceNickName
            ws[f"B{rowMax}"].number_format = "@"

            ws[f"C{rowMax}"].value = workpieceDimension
            ws[f"C{rowMax}"].number_format = "@"
            ws[f"D{rowMax}"].value = workpieceMaterial
            ws[f"D{rowMax}"].number_format = "@"
            ws[f"E{rowMax}"].value = workpiece1stParameter
            ws[f"E{rowMax}"].number_format = "@"
            if not workpiece2ndParameter or not re.search(r"^\d", workpiece2ndParameter):
                ws[f"F{rowMax}"].value = workpiece2ndParameter
                ws[f"F{rowMax}"].number_format = "@"
                # DEPRECATED:
                # ws[f"H{rowMax}"].value = workpiece2ndParameterNum
                # ws[f"H{rowMax}"].number_format = BUILTIN_FORMATS[2]
            ws[f"G{rowMax}"].value = workpieceLength
            ws[f"G{rowMax}"].number_format = "0.0"

        # Calculate the surface area
        if workpieceDimension and fileNameMatchTick and "∅" in workpieceDimension and "L" in workpieceDimension:
            m = cfg.patterns.workpieceDimension.match(workpieceDimension)
            if m:
                dia    = float(m.group(1)[1:])
                length = float(m.group(3)[1:])
                surfaceAreaFormula = f"=3.14 * { dia } * G{rowMax} / 1000 / 1000"
                surfaceAreaEval = 3.14 * dia * length / 1000 / 1000
                ws[f"H{rowMax}"].value = surfaceAreaFormula
                ws[f"H{rowMax}"].number_format = "0.0000"

        # Use override area
        areaOverride = workpieceDict["areaOverride"]
        if workpieceFullName in areaOverride or (workpieceNickName and workpieceNickName in areaOverride):
            if workpieceNickName:
                querryKey = workpieceNickName
            else:
                querryKey = workpieceFullName
            overrideVal = areaOverride[querryKey]

            if isinstance(overrideVal, float):
                if ws[f"H{rowMax}"].value and surfaceAreaEval:
                    pr(f"Override area for {querryKey} with {areaOverride[querryKey]} instead of {surfaceAreaEval}")
                else:
                    pr(f"Override area for {querryKey} with {areaOverride[querryKey]}")

                ws[f"H{rowMax}"].value = areaOverride[querryKey]
            elif isinstance(overrideVal, list):
                ws[f"I{rowMax}"].value = "\n".join(overrideVal)
                ws[f"I{rowMax}"].number_format = "@"
                ws[f"H{rowMax}"].value = f"=SUMPRODUCT(SUMIF($B:$B,TEXTSPLIT($I{rowMax},CHAR(10)),$H:$H))+SUMPRODUCT(SUMIF($A:$A,TEXTSPLIT($I{rowMax},CHAR(10)),$H:$H))"
            elif isinstance(overrideVal, str):
                ws[f"H{rowMax}"].value = f'=IF(ISNUMBER(MATCH("{overrideVal}", B:B, 0)), INDEX(H:H, MATCH("{overrideVal}", B:B, 0)), IF(ISNUMBER(MATCH("{overrideVal}", A:A, 0)), INDEX(H:H, MATCH("{overrideVal}", A:A, 0)), ""))'
                pr(f"area of {querryKey} is linked to {areaOverride[querryKey]}")

            ws[f"H{rowMax}"].number_format = "0.0000"

    # Add table
    tab = Table(displayName="Table1", ref=f"A2:I{ws.max_row}")

    # Add printable area
    ws.oddFooter.center.text = "第 &[Page] 页，共 &N 页" # type: ignore
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_title_rows = "2:2"
    ws.print_area = f"A2:I{ws.max_row}"

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
            name="TableStyleMedium16",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
            )
    tab.tableStyleInfo = style

    ws.add_table(tab)

    # Add protection
    # ws.protection.sheet = True
    # ws.protection.sort = False
    # ws.protection.autoFilter = False
    # ws.protection.password = '456'
    # ws.protection.enable()

    if dstPath1.exists():
        savePath = util.saveWorkbook(wb, dstPath1, True)
    if dstPath2.exists():
        savePath = util.saveWorkbook(wb, dstPath2, False)
