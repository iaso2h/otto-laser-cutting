import util
from config import cfg
import keySet
import style

import chardet
import os
import re
import datetime
from typing import Optional
from collections import Counter
from pathlib import Path
from striprtf.striprtf import rtf_to_text
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Protection, Alignment
from pprint import pprint

pr = util.pr
LASER_PROFILE_PATH = Path(cfg.paths.otto, r"存档/耗时计算.xlsx")
TUBEPRO_LOG_PATH   = Path(cfg.paths.otto, r"存档/切割机日志")
fileOpenPat      = re.compile(r"^\(([0-9:\/ ]+?)\)打开文件：(.+)")
segmentFirstPat  = re.compile(r"^\(([0-9:\/ ]+?)\)总零件数:(\d+), 当前零件序号:1$")
segmentPat       = re.compile(r"^\(([0-9:\/ ]+?)\)总零件数:(\d+), 当前零件序号:(\d+)")
scheduelTotalPat = re.compile(r"^\(([0-9:\/ ]+?)\).+零件切割计划数目(\d+)")
loopEndPat       = re.compile(r"^\(([0-9:\/ ]+?)\).+已切割零件数目(\d+)")
# loopStartPat     = re.compile(r"^\(([0-9:\/ ]+?)\)开始加工.{1,3}循环计数：(\d+)")


def getEncoding(filePath) -> str:
    # Create a magic object
    """
    Detects the encoding of a file using chardet.

    Args:
        filePath (str): Path to the file to analyze.

    Returns:
        str: Detected encoding as a string (e.g. 'utf-8'), or empty string if detection fails.
    """
    with open(filePath, "rb") as f:
        # Detect the encoding
        rawData = f.read()
        result = chardet.detect(rawData)
        if not result:
            return ""
        if not result["encoding"]:
            return ""
        else:
            return result["encoding"]


def fillWorkbook(ws: Worksheet, parsedResult: dict, sortChk: bool):
    """
    Fills an Excel worksheet with laser cutting file statistics from parsed data.

    Args:
        ws (Worksheet): OpenPyXL Worksheet object to populate with data.
        parsedResult (dict): Dictionary containing parsed laser file information with keys:
            - loop: List of loop intervals
            - loopIntervalCounter: Counter object of interval frequencies
            - loopIntervalUpdated: Dictionary of last update timestamps per interval
            - workpieceCount: Number of workpieces per file
        sortChk (bool): Whether to sort the results alphabetically by filename.

    Populates worksheet with:
        - Headers in row 1 with formatted columns
        - File statistics grouped by laser filename
        - Calculated fields for material/time consumption
        - Cell protection with password '456'
    """
    ws[f"A{1}"].value = "排样文件"
    ws.column_dimensions["A"].width = 35
    ws[f"B{1}"].value = "循环耗时"
    ws.column_dimensions["B"].width = 12
    ws[f"C{1}"].value = "循环统计"
    ws.column_dimensions["C"].width = 12
    ws[f"D{1}"].value = "最后统计日期"
    ws.column_dimensions["D"].width = 22
    ws[f"E{1}"].value = "工件目标数"
    ws.column_dimensions["E"].width = 14
    ws[f"F{1}"].value = "工件已加工数"
    ws.column_dimensions["F"].width = 17
    ws[f"G{1}"].value = "预计消耗长料"
    ws.column_dimensions["G"].width = 17
    ws[f"H{1}"].value = "预计消耗时长"
    ws.column_dimensions["H"].width = 17
    ws[f"I{1}"].value = "预计完成时间"
    ws.column_dimensions["I"].width = 22
    for col in range(1, 10):
        ws.cell(row=1, column=col).style     = "Headline 1"
        ws.cell(row=1, column=col).alignment = style.alCenter

    if sortChk:
        items = sorted(parsedResult.items())
    else:
        items = parsedResult.items()
    for laserFileName, laserFileInfo in items:
        if len(laserFileInfo["loop"]) < 1:
            continue

        mostCommon = laserFileInfo["loopIntervalCounter"].most_common(5)
        laserFileStartRow = ws.max_row
        skipRowCount = 0
        headlineBorderSet = False
        for intervalIdx, common in enumerate(mostCommon):
            currentRow = intervalIdx + laserFileStartRow + 1 - skipRowCount
            interval      = common[0]
            intervalCount = common[1]

            if intervalIdx == len(mostCommon) - 1:

                # Merge laser filename cells under column A
                if interval == "0":
                    endRow = currentRow - 1
                else:
                    endRow = currentRow
                # Check merge necessity of merging cells
                if endRow > laserFileStartRow + 1:
                    ws.merge_cells(
                        start_row    = laserFileStartRow + 1,
                        end_row      = endRow,
                        start_column = 1,
                        end_column   = 1
                    )
                    ws.cell(row=laserFileStartRow+1,column=1).alignment = style.alCenterWrap

            # Don't fill in sheet when interval between two loop is 0
            if interval == "0":
                skipRowCount += 1
                continue

            ws.cell(row=currentRow, column=2).value = int(float(interval))
            ws.cell(row=currentRow, column=2).number_format = '0"秒"'
            ws.cell(row=currentRow, column=3).value = intervalCount
            ws.cell(row=currentRow, column=3).number_format = '0"次"'
            ws.cell(row=currentRow, column=4).value = laserFileInfo["loopIntervalUpdated"][interval]
            ws.cell(row=currentRow, column=5).value = 100
            ws.cell(row=currentRow, column=5).font = style.font["orangeBold"]
            ws.cell(row=currentRow, column=5).number_format = '0"支"'
            ws.cell(row=currentRow, column=5).protection = Protection(locked=False)
            ws.cell(row=currentRow, column=6).value = 0
            ws.cell(row=currentRow, column=6).font = style.font["greenBold"]
            ws.cell(row=currentRow, column=6).number_format = '0"支"'
            ws.cell(row=currentRow, column=6).protection = Protection(locked=False)
            ws.cell(row=currentRow, column=7).value = f'=(E{currentRow}-F{currentRow})/{laserFileInfo["workpieceCount"]}'
            ws.cell(row=currentRow, column=7).number_format = '0"支"'
            ws.cell(row=currentRow, column=8).value = f'=(B{currentRow}+1)/{laserFileInfo["workpieceCount"]}*(E{currentRow}-F{currentRow})/86400'
            ws.cell(row=currentRow, column=8).number_format = "[h]时mm分ss秒"
            ws.cell(row=currentRow, column=9).value = f'=NOW() + H{currentRow}'
            ws.cell(row=currentRow, column=9).number_format = "yyyy-m-d h:mm:ss"

            if not headlineBorderSet:
                # Add top border
                headlineBorderSet = True
                ws.cell(row=currentRow, column=1).value = laserFileName
                ws[f"A{currentRow}"].border = style.borderMedium
                ws[f"B{currentRow}"].border = style.borderMedium
                ws[f"C{currentRow}"].border = style.borderMedium
                ws[f"D{currentRow}"].border = style.borderMedium
                ws[f"E{currentRow}"].border = style.borderMedium
                ws[f"F{currentRow}"].border = style.borderMedium
                ws[f"G{currentRow}"].border = style.borderMedium
                ws[f"H{currentRow}"].border = style.borderMedium
                ws[f"I{currentRow}"].border = style.borderMedium

            ws.protection.sheet = True
            ws.protection.password = '456'
            ws.protection.enable()


def parse(
    rtfFile: Path,
    wb: Workbook,
    accumulationMode: bool,
    parsedResult: Optional[dict] = None,
) -> dict:
    """
    Parses an RTF file containing laser cutting records and organizes the data into a structured format.

    Args:
        rtfFile: Path to the RTF file to parse
        wb: Workbook object for Excel output (optional, used in non-accumulation mode)
        accumulationMode: If True, accumulates results without writing to Excel
        parsedResult: Optional dictionary to accumulate results across multiple files

    Returns:
        Dictionary containing:
        - "workbook": Modified Workbook object (None in accumulation mode)
        - "parsedResult": Dictionary of parsed data with structure:
            {
                "laserFileName": {
                    "open": [(lineIdx, timestamp)],
                    "loop": [(lineIdx, timestamp, interval)],
                    "loopIntervalUpdated": {interval: timestamp},
                    "loopIntervalCounter": Counter(intervals),
                    "workpieceCount": int
                }
            }

    The function processes RTF content to extract:
    1. Laser file open events with timestamps
    2. Loop segments with timestamps and intervals
    3. Workpiece counts
    4. Statistics on loop intervals
    """
    laserFileLastOpen = ""
    loopLastTime = None
    if parsedResult is None:
        parsedResult = {}

    now = datetime.datetime.now()
    with open(rtfFile, "r", encoding=getEncoding(str(rtfFile))) as f:
        content = rtf_to_text(f.read())
        lines = content.split("\n")

    laserFileFullPath = ""
    for lineIdx, line in enumerate(lines):
        fileOpenMatch  = fileOpenPat.match(line)
        loopStartMatch = segmentFirstPat.match(line)
        if fileOpenMatch:
            laserFileFullPath = fileOpenMatch.group(2)
            laserFileName = laserFileFullPath.replace("D:\\欧拓图纸\\切割文件\\", "")
            laserFileName = util.diametartSymbolUnify(laserFileName)
            laserFileName = laserFileName.replace(".zx", ".zzx")
            laserFileName = laserFileName.replace("  ", " ")
            laserFileName = laserFileName.strip()
            laserFileLastOpen = laserFileName
            if laserFileName not in parsedResult:
                parsedResult[laserFileName] = {
                    "open": [],
                    "loop": [],
                    "loopIntervalUpdated": {},
                    "loopIntervalCounter": Counter(),
                    "workpieceCount": 0
                }
                loopLastTime = None
            parsedResult[laserFileName]["open"].append(( lineIdx, fileOpenMatch.group(1) ))


        if loopStartMatch:
            timeStamp = loopStartMatch.group(1)
            timeLoop  = datetime.datetime.strptime(f"{now.year}/{timeStamp}", "%Y/%m/%d %H:%M:%S")
            if not loopLastTime:
                loopInterval = 0
            else:
                loopInterval = (timeLoop - loopLastTime).total_seconds()

            # Add addiontional time window in accumulation mode
            if accumulationMode:
                if loopInterval:
                    loopInterval += 15

            loopLastTime = timeLoop

            parsedResult[laserFileLastOpen]["loop"].append(( lineIdx, timeStamp, loopInterval))
            parsedResult[laserFileLastOpen]["loopIntervalUpdated"][f"{loopInterval}"] = loopLastTime
            parsedResult[laserFileLastOpen]["loopIntervalCounter"][f"{loopInterval}"] += 1
            # Get maximun workpiece count
            if int(loopStartMatch.group(2)) > parsedResult[laserFileLastOpen]["workpieceCount"]:
                parsedResult[laserFileLastOpen]["workpieceCount"] = int(loopStartMatch.group(2))

    if not parsedResult:
        pr(f"No laser file records parsed from rtf file {str(rtfFile)}")
        return {
            "workbook":     None,
            "parsedResult": None
        }

    if accumulationMode:
        return {
                "workbook":     None,
                "parsedResult": parsedResult
                }
    else:
        if wb.active.title == "Sheet": # type: ignore
            ws = wb.active # type: ignore
            ws.title = rtfFile.stem # type: ignore
        else:
            ws = wb.create_sheet(rtfFile.stem, 0)
        fillWorkbook(ws, parsedResult, True) # type: ignore
        return {
                "workbook":     wb,
                "parsedResult": parsedResult
                }


def parseAccuLog():
    """
    Parses accumulated laser cutting logs from RTF files within the last 60 days.
    Processes files in TUBEPRO_LOG_PATH (excluding '精简' files), extracts data into a workbook,
    and saves results to LASER_PROFILE_PATH. Skips hidden column F in output.

    Returns:
        None: Prints message if no logs found, otherwise saves processed data to file.
    """
    wb = Workbook()
    parsedResult = None
    now = datetime.datetime.now()
    timeDeltaLiteral = 60
    timeDelta = datetime.timedelta(days=timeDeltaLiteral)

    for f in Path(TUBEPRO_LOG_PATH).iterdir():
        if f.suffix != ".rtf":
            continue

        logTime = datetime.datetime.fromtimestamp(f.stat().st_ctime)
        if now - logTime <= timeDelta:
            parsedResult = parse(
                rtfFile=f,
                wb=wb,
                accumulationMode=True,
                parsedResult=parsedResult
                    )["parsedResult"] # type: ignore
    if not parsedResult:
        return pr("No parsed accumulated result")

    fillWorkbook(wb.active, parsedResult, True) # type: ignore
    wb.active.column_dimensions['F'].hidden = True #type: ignore
    util.saveWorkbook(wb, LASER_PROFILE_PATH, True) # type: ignore


def parsePeriodLog():
    """
    Parses laser cutting log files within a specified time period based on modifier keys.
    Handles different parsing modes:
    - Ctrl+Shift+Alt: Parse accumulated logs (calls parseAccuLog)
    - Ctrl: Open laser profile directly
    - Shift: Parse logs from last 7 days
    - Alt: Parse all logs (calls parseAllLog)
    - No modifier: Parse logs from last 1 day
    Automatically expands time window (up to 3 attempts) if no logs found.
    Saves parsed data to laser profile if logs were processed.
    """
    if "ctrl" in keySet.keys and "shift" in keySet.keys and "alt" in keySet.keys:
        return parseAccuLog()
    elif "ctrl" in keySet.keys:
        return os.startfile(LASER_PROFILE_PATH)
    elif "shift" in keySet.keys:
        timeDeltaLiteral = 7
    elif "alt" in keySet.keys:
        timeDeltaLiteral = 365
    else:
        timeDeltaLiteral = 1

    wb = Workbook()
    now = datetime.datetime.now()
    parsedPeriodCount = 0
    for loopCount in range(3):
        if parsedPeriodCount > 0:
            break
        else:
            # Increase the time delta window to if no parsed files
            timeDeltaLiteral = timeDeltaLiteral * (7 ** loopCount)
        timeDelta = datetime.timedelta(days=timeDeltaLiteral)

        for f in Path(TUBEPRO_LOG_PATH).iterdir():
            if f.suffix != ".rtf":
                continue

            logTime = datetime.datetime.fromtimestamp(f.stat().st_ctime)
            if now - logTime <= timeDelta:
                wb = parse(
                    rtfFile=f,
                    wb=wb,
                    accumulationMode=False
                    )["workbook"] # type: ignore
                parsedPeriodCount += 1

    if parsedPeriodCount:
        util.saveWorkbook(wb, LASER_PROFILE_PATH, True) # type: ignore


def rtfSimplify():
    """
    Processes RTF log files in TUBEPRO_LOG_PATH based on modifier keys:
    - Ctrl: Opens the log directory
    - Shift: Processes files from last 7 days
    - Alt: Processes files from last year
    - No modifier: Processes files from last day

    For each matching RTF file:
    1. Filters content using regex patterns (laserFileOpenPat, segmentPat, etc.)
    2. Creates a simplified version with '精简' prefix in filename
    3. Outputs processed files with relevant log lines

    Handles cases where no files are found by expanding time window exponentially.
    """
    if "ctrl" in keySet.keys:
        return os.startfile(TUBEPRO_LOG_PATH)
    elif "shift" in keySet.keys:
        timeDeltaLiteral = 7
    elif "alt" in keySet.keys:
        timeDeltaLiteral = 365
    else:
        timeDeltaLiteral = 1


    parsedPeriodCount = 0
    cuttingSessions = []
    for loopCount in range(3):
        if parsedPeriodCount > 0:
            break
        else:
            # Increase the time delta window to if no parsed files
            timeDeltaLiteral = timeDeltaLiteral * (7 ** loopCount)
        timeDelta = datetime.timedelta(days=timeDeltaLiteral)

        # Iterating through all rtf files
        for rtfFile in TUBEPRO_LOG_PATH.glob("*.rtf"):
            now = datetime.datetime.now()
            rtfCreationTime = datetime.datetime.fromtimestamp(rtfFile.stat().st_ctime)
            if now - rtfCreationTime > timeDelta:
                continue
            with open(rtfFile, "r", encoding=getEncoding(str(rtfFile))) as f:
                rtfContent = f.read()
                content = rtf_to_text(rtfContent)
                lines = content.split("\n")

            # Itering through all lines in rtf file to filter out the lines
            refinedLines = []
            for line in lines:
                fileOpenMatch      = fileOpenPat.match(line)
                segmentMatch       = segmentPat.match(line)
                scheduelTotalMatch = scheduelTotalPat.match(line)
                loopEndMatch       = loopEndPat.match(line)
                if not any((
                    fileOpenMatch,
                    segmentMatch,
                    scheduelTotalMatch,
                    loopEndMatch,
                )):
                    continue

                if fileOpenMatch:
                    timeStamp = fileOpenMatch.group(1)
                    currentFileOpen = Path(fileOpenMatch.group(2)).stem
                    currentFileOpen = currentFileOpen.replace("_X1", "")
                    timeObj = datetime.datetime.strptime(
                        f"{now.year}/{timeStamp}",
                        "%Y/%m/%d %H:%M:%S"
                    )
                    if not cuttingSessions or cuttingSessions[len(cuttingSessions) - 1]["fileName"]["value"] != currentFileOpen:
                        cuttingSession = {
                            "fileName":       {"value": currentFileOpen, "updatedTime": timeObj },
                            "startCount":     {"value": -1, "updatedTime": timeObj },
                            "segmentTotal":   {"value": 0, "updatedTime": None},
                            "segmentCount":   {"value": 0, "updatedTime": None},
                            "scheduleTotal":  {"value": 0, "updatedTime": None},
                            "loopEndCount":   {"value": 0, "updatedTime": None},
                            "totalCount":     {"value": 0, "updatedTime": None},
                        }
                        cuttingSessions.append(cuttingSession)


                if segmentMatch:
                    timeStamp = segmentMatch.group(1)
                    timeObj = datetime.datetime.strptime(
                        f"{now.year}/{timeStamp}",
                        "%Y/%m/%d %H:%M:%S"
                    )
                    cuttingSessions[len(cuttingSessions) - 1]["segmentTotal"] = {
                        "value": int(segmentMatch.group(2)),
                        "updatedTime": timeObj
                    }
                    cuttingSessions[len(cuttingSessions) - 1]["segmentCount"] = {
                        "value": int(segmentMatch.group(3)),
                        "updatedTime": timeObj
                    }
                if scheduelTotalMatch:
                    timeStamp = scheduelTotalMatch.group(1)
                    timeObj = datetime.datetime.strptime(
                        f"{now.year}/{timeStamp}",
                        "%Y/%m/%d %H:%M:%S"
                    )
                    cuttingSessions[len(cuttingSessions) - 1]["scheduleTotal"] = {
                        "value": int(scheduelTotalMatch.group(2)),
                        "updatedTime": timeObj
                    }
                if loopEndMatch:
                    timeStamp = loopEndMatch.group(1)
                    timeObj = datetime.datetime.strptime(
                        f"{now.year}/{timeStamp}",
                        "%Y/%m/%d %H:%M:%S"
                    )
                    cuttingSessions[len(cuttingSessions) - 1]["loopEndCount"] = {
                        "value": int(loopEndMatch.group(2)),
                        "updatedTime": timeObj
                    }
                    # determin whether the cutting session was starting from 0
                    if cuttingSessions[len(cuttingSessions) - 1]["startCount"]["value"] == -1:
                        if cuttingSessions[len(cuttingSessions) - 1]["loopEndCount"]["value"] == cuttingSessions[len(cuttingSessions) - 1]["segmentTotal"]["value"]:
                            cuttingSessions[len(cuttingSessions) - 1]["startCount"]["value"] = 0
                        else:
                            cuttingSessions[len(cuttingSessions) - 1]["startCount"]["value"] = (
                                cuttingSessions[len(cuttingSessions) - 1]["loopEndCount"]["value"]
                                - cuttingSessions[len(cuttingSessions) - 1]["segmentTotal"]["value"]
                            )
                            if cuttingSessions[len(cuttingSessions) - 1]["startCount"]["value"] < 0:
                                cuttingSessions[len(cuttingSessions) - 1]["startCount"]["value"] = 0



                refinedLines.append(line + "\n")
            targetPath = Path(
                rtfFile.parent,
                "精简",
                rtfFile.stem + rtfFile.suffix
            )
            os.makedirs(targetPath.parent, exist_ok=True)
            try:
                with open(targetPath, mode="w", encoding="utf-8") as f2:
                    for line in refinedLines:
                        f2.write(line)
                finishTime = datetime.datetime.now()
                delta = finishTime - now
                pr(f"导出文件: {str(rtfFile)}，耗时: {delta.total_seconds()}秒")
                parsedPeriodCount += 1
            except PermissionError:
                pr(f"无法写入文件: {str(targetPath)}，请检查文件是否被占用或权限设置。")

    if cuttingSessions:
        pr("rtf日志精简完成")
    else:
        return pr("没有rtf日志被分析")

    # Compute the total count of for every file opened in every session
    for session in cuttingSessions:
        if session["loopEndCount"]["updatedTime"] and session["loopEndCount"]["updatedTime"] > session["segmentCount"]["updatedTime"]:
            session["totalCount"]["value"] = session["loopEndCount"]["value"]
            session["totalCount"]["updatedTime"] = session["loopEndCount"]["updatedTime"]
        else:
            session["totalCount"]["value"] = session["segmentCount"]["value"] + session["loopEndCount"]["value"]
            session["totalCount"]["updatedTime"] = session["segmentCount"]["updatedTime"]

        # Rectify the start count if the cutting session wasn't starting from 0
        if session["startCount"]["value"] == -1 and session["loopEndCount"]["value"] == 0:
            session["startCount"]["value"] = 0


    # Exporting cuttingSessions data into Excel file
    wb = Workbook()
    ws = wb.active # type: Worksheet
    ws["A1"].value = "排样名称"
    ws.column_dimensions["A"].width = 50
    ws["B1"].value = "开始时间"
    ws.column_dimensions["B"].width = 22
    ws["C1"].value = "结束时间"
    ws.column_dimensions["C"].width = 22
    ws["D1"].value = "耗时"
    ws.column_dimensions["D"].width = 12.5
    ws["E1"].value = "开始数量"
    ws.column_dimensions["E"].width = 10
    ws["F1"].value = "结束数量"
    ws.column_dimensions["F"].width = 10
    ws["G1"].value = "目标数量"
    ws.column_dimensions["G"].width = 10
    skipCount = 0
    for sessionIdx, session in enumerate(cuttingSessions):
        if session["totalCount"]["value"] == 0:
            skipCount += 1
            continue
        row = 2 + sessionIdx - skipCount
        ws[f"A{row}"].value = session["fileName"]["value"]
        ws[f"A{row}"].alignment = Alignment(wrapText = True)
        ws[f"B{row}"].value = session["fileName"]["updatedTime"].strftime("%Y/%m/%d %H:%M:%S")
        ws[f"B{row}"].number_format = "yyyy/m/d h:mm:ss"
        ws[f"C{row}"].value = session["totalCount"]["updatedTime"].strftime("%Y/%m/%d %H:%M:%S")
        ws[f"C{row}"].number_format = "yyyy/m/d h:mm:ss"
        ws[f"D{row}"].value = f'=C{row}-B{row}'
        ws[f"D{row}"].number_format = "[h]时mm分ss秒"
        ws[f"E{row}"].value = session["startCount"]["value"]
        ws[f"F{row}"].value = session["totalCount"]["value"]
        ws[f"G{row}"].value = session["scheduleTotal"]["value"]


    # Add table
    tab = Table(displayName="Table1", ref=f"A1:G{ws.max_row}")

    # Add printable area
    ws.oddFooter.center.text = "第 &[Page] 页，共 &N 页" # type: ignore
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_title_rows = "2:2"
    ws.print_area = f"A1:G{ws.max_row}"

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
            name="TableStyleMedium24",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
            )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    util.saveWorkbook(wb)


